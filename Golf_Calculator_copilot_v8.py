import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

MAX_PLAYERS = 40
MAX_HOLE_SCORE = 9
HOLES = 18

# canonical reason codes for hole results
REASON_NO_SCORES = "NO_SCORES"
REASON_SOLE = "SOLE"
REASON_CARRY = "CARRY"
REASON_SPLIT = "SPLIT"


class PlayerRow:
	def __init__(self, parent, idx, app):
		# place widgets directly into the parent grid so their columns align with header labels
		self.parent = parent
		# reference to parent application (to read pars for highlighting)
		self.app = app
		self.row = idx + 1  # header occupies row 0
		self.name_var = tk.StringVar()
		self.handicap_var = tk.StringVar(value="0")
		self.include_var = tk.BooleanVar(value=True)
		self.score_vars = [tk.StringVar(value="") for _ in range(HOLES)]
		# keep entry widgets so we can color them for birdie/eagle
		self.score_entries = []
		self.score_entry_defaults = []

		# create widgets directly on parent at proper columns
		self.name_entry = ttk.Entry(self.parent, textvariable=self.name_var, width=20)
		self.name_entry.grid(row=self.row, column=0, padx=2, pady=2)

		self.handicap_entry = ttk.Entry(self.parent, textvariable=self.handicap_var, width=5)
		self.handicap_entry.grid(row=self.row, column=1, padx=2, pady=2)

		self.include_cb = ttk.Checkbutton(self.parent, variable=self.include_var)
		self.include_cb.grid(row=self.row, column=2, padx=2, pady=2)

		for h in range(HOLES):
			# use tk.Entry for easy background color changes across platforms
			e = tk.Entry(self.parent, textvariable=self.score_vars[h], width=3)
			e.grid(row=self.row, column=3 + h, padx=1, pady=2)
			# remember default bg for later restore
			try:
				self.score_entry_defaults.append(e.cget('bg'))
			except Exception:
				self.score_entry_defaults.append(None)
			self.score_entries.append(e)
			e.bind("<FocusOut>", lambda ev, sv=self.score_vars[h]: self._validate_score_var(sv))

		self.front9_lbl = ttk.Label(self.parent, text="0", width=5)
		self.front9_lbl.grid(row=self.row, column=3 + HOLES, padx=4)
		self.back9_lbl = ttk.Label(self.parent, text="0", width=5)
		self.back9_lbl.grid(row=self.row, column=4 + HOLES, padx=4)

		for sv in self.score_vars:
			sv.trace_add("write", lambda *a: self.update_totals())

	def _validate_score_var(self, var):
		v = var.get().strip()
		if v == "":
			return
		try:
			n = int(v)
			if n < 0 or n > MAX_HOLE_SCORE:
				messagebox.showwarning("Invalid score", f"Score must be between 0 and {MAX_HOLE_SCORE}")
				var.set("")
		except ValueError:
			messagebox.showwarning("Invalid score", "Score must be an integer")
			var.set("")

	def update_totals(self):
		front = 0
		back = 0
		for i in range(9):
			v = self.score_vars[i].get().strip()
			if v.isdigit():
				front += int(v)
		for i in range(9, 18):
			v = self.score_vars[i].get().strip()
			if v.isdigit():
				back += int(v)
		self.front9_lbl.config(text=str(front))
		self.back9_lbl.config(text=str(back))

		# Highlight birdies/eagles on the per-hole entry widgets if possible
		for i in range(HOLES):
			try:
				ent = self.score_entries[i]
			except Exception:
				continue
			val = self.score_vars[i].get().strip()
			# reset to default if empty or non-numeric
			if val == "":
				try:
					default = self.score_entry_defaults[i]
					if default is not None:
						ent.config(bg=default)
					else:
						ent.config(bg='white')
				except Exception:
					pass
				continue
			try:
				score = int(val)
			except Exception:
				# non-numeric -> leave default
				continue
			# read par for this hole from app (fall back to 4)
			try:
				par_v = int(self.app.par_vars[i].get())
			except Exception:
				par_v = 4
			# colors chosen to match export highlights
			if score == par_v - 1:
				try:
					ent.config(bg="#FFF59D")
				except Exception:
					pass
			elif score <= par_v - 2:
				try:
					ent.config(bg="#C8E6C9")
				except Exception:
					pass
			else:
				try:
					default = self.score_entry_defaults[i]
					if default is not None:
						ent.config(bg=default)
					else:
						ent.config(bg='white')
				except Exception:
					pass

	def destroy(self):
		# destroy widgets we created
		try:
			self.name_entry.destroy()
		except:
			pass
		try:
			self.handicap_entry.destroy()
		except:
			pass
		try:
			self.include_cb.destroy()
		except:
			pass
		for h in range(HOLES):
			try:
				# find widget at the grid position and destroy if exists
				w = self.parent.grid_slaves(row=self.row, column=3 + h)
				if w:
					w[0].destroy()
			except:
				pass
		try:
			self.front9_lbl.destroy()
		except:
			pass
		try:
			self.back9_lbl.destroy()
		except:
			pass

	def to_dict(self):
		d = {
			"Name": self.name_var.get().strip(),
			"Handicap": self.handicap_var.get().strip(),
			"Included": self.include_var.get()
		}
		for i in range(HOLES):
			d[f"H{i+1}"] = self.score_vars[i].get().strip()
		d["Front9"] = self.front9_lbl.cget("text")
		d["Back9"] = self.back9_lbl.cget("text")
		return d

	def load_from_dict(self, d):
		self.name_var.set(d.get("Name", ""))
		self.handicap_var.set(str(d.get("Handicap", "0")))
		self.include_var.set(bool(d.get("Included", True)))
		for i in range(HOLES):
			key = f"H{i+1}"
			val = d.get(key, "")
			# pandas may use NaN for missing; guard with pd.isna
			if pd.isna(val):
				self.score_vars[i].set("")
			else:
				self.score_vars[i].set(str(val))
		self.update_totals()


class BigBoySkinsApp:
	def __init__(self, root):
		self.root = root
		self.root.title("Big Boy Skins Manager")
		self.players = []
		self.use_net_scores = tk.BooleanVar(value=False)
		self.split_ties = tk.BooleanVar(value=False)
		self.course_name_var = tk.StringVar()
		# backward-compatible alias used by export/import code
		self.course_var = self.course_name_var
		self.date_var = tk.StringVar()
		# export/import controls
		self.per_skin_var = tk.StringVar(value="1")
		self.total_purse_var = tk.StringVar(value="")
		self.carryover_var = tk.BooleanVar(value=True)
		# enable/disable birdie/eagle bonus skins (birdie=+1, eagle=+5)
		# default to enabled
		self.bonus_enabled_var = tk.BooleanVar(value=True)
		self.par_vars = [tk.StringVar(value="4") for _ in range(HOLES)]
		self.stroke_index_vars = [tk.StringVar(value=str(i+1)) for i in range(HOLES)]

		self.build_gui()

	def build_gui(self):
		header = ttk.Frame(self.root)
		header.grid(row=0, column=0, sticky="w", padx=10, pady=5)

		# allow the main window and header to expand
		self.root.columnconfigure(0, weight=1)
		header.columnconfigure(1, weight=2)  # course name entry
		header.columnconfigure(3, weight=1)  # date entry
		header.columnconfigure(5, weight=1)  # per-skin
		header.columnconfigure(7, weight=1)  # total purse

		ttk.Label(header, text="Course Name:").grid(row=0, column=0)
		ttk.Entry(header, textvariable=self.course_name_var, width=30).grid(row=0, column=1, sticky="ew")

		ttk.Label(header, text="Date:").grid(row=0, column=2)
		self.date_entry = DateEntry(header, textvariable=self.date_var, width=12)
		self.date_entry.grid(row=0, column=3, sticky="ew")
		# export/import controls (per-skin vs total purse and carryover)
		ttk.Label(header, text="Per Skin $").grid(row=0, column=4)
		ttk.Entry(header, textvariable=self.per_skin_var, width=6).grid(row=0, column=5, sticky="ew")
		ttk.Label(header, text="Total Purse $").grid(row=0, column=6)
		tk.Entry(header, textvariable=self.total_purse_var, width=8).grid(row=0, column=7, sticky="ew")
		tk.Checkbutton(header, text='Carryover', variable=self.carryover_var).grid(row=0, column=8)
		# Option to enable birdie/eagle bonus skins (birdie +1, eagle +5)
		# When enabled, a sole winner at par-1 gains +1 unit, at <=par-2 gains +5 units
		tk.Checkbutton(header, text='Apply Birdie/Eagle Bonuses', variable=self.bonus_enabled_var).grid(row=0, column=9)

		ttk.Checkbutton(header, text="Use Net Scores (based on handicap)", variable=self.use_net_scores).grid(row=1, column=0, columnspan=2, sticky="w")
		tk.Checkbutton(header, text="Split Ties (instead of carryover)", variable=self.split_ties).grid(row=1, column=2, columnspan=2, sticky="w")

		# Par and Stroke Index rows
		setup_frame = ttk.Frame(self.root)
		setup_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=5)
		# make par/stroke index entries expand
		for ci in range(1, 1 + HOLES):
			setup_frame.columnconfigure(ci, weight=1)

		ttk.Label(setup_frame, text="Par:").grid(row=0, column=0)
		for i in range(HOLES):
			tk.Entry(setup_frame, textvariable=self.par_vars[i], width=3).grid(row=0, column=i+1)

		ttk.Label(setup_frame, text="Stroke Index:").grid(row=1, column=0)
		for i in range(HOLES):
			tk.Entry(setup_frame, textvariable=self.stroke_index_vars[i], width=3).grid(row=1, column=i+1)

		# Player entry frame
		self.player_frame = ttk.Frame(self.root)
		self.player_frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
		# allow the player grid columns to expand; make Name column a bit wider
		self.player_frame.columnconfigure(0, weight=2)
		for ci in range(1, 8 + HOLES):
			self.player_frame.columnconfigure(ci, weight=1)

		ttk.Label(self.player_frame, text="Name").grid(row=0, column=0)
		ttk.Label(self.player_frame, text="HCP").grid(row=0, column=1)
		ttk.Label(self.player_frame, text="In?").grid(row=0, column=2)
		for i in range(HOLES):
			ttk.Label(self.player_frame, text=f"H{i+1}").grid(row=0, column=3 + i)
		tk.Label(self.player_frame, text="Front9").grid(row=0, column=3 + HOLES)
		tk.Label(self.player_frame, text="Back9").grid(row=0, column=4 + HOLES)

		# Buttons
		btn_frame = ttk.Frame(self.root)
		btn_frame.grid(row=3, column=0, sticky="w", padx=10, pady=10)

		ttk.Button(btn_frame, text="Add Player", command=self.add_player).grid(row=0, column=0, padx=5)
		ttk.Button(btn_frame, text="Export to Excel", command=self.export_to_excel).grid(row=0, column=1, padx=5)
		ttk.Button(btn_frame, text="Import from Excel", command=self.import_from_excel).grid(row=0, column=2, padx=5)

		# Add initial players
		for _ in range(2):
			self.add_player()

	def add_player(self):
		if len(self.players) >= MAX_PLAYERS:
			messagebox.showwarning("Limit reached", f"Maximum {MAX_PLAYERS} players allowed.")
			return
		idx = len(self.players)
		row = PlayerRow(self.player_frame, idx, self)
		# widgets are created and gridded by PlayerRow directly
		self.players.append(row)

	def collect_data(self):
		pars = []
		for v in self.par_vars:
			s = v.get().strip()
			pars.append(int(s) if s.isdigit() else 4)
		player_dicts = []
		for pr in self.players:
			d = pr.to_dict()
			if d.get("Name", "") == "":
				continue
			for i in range(HOLES):
				key = f"H{i+1}"
				val = d.get(key, "")
				if isinstance(val, str) and val.strip() == "":
					d[key] = ""
				else:
					try:
						n = int(val)
						if n < 0 or n > MAX_HOLE_SCORE:
							d[key] = ""
						else:
							d[key] = n
					except Exception:
						d[key] = ""
			player_dicts.append(d)
		return pars, player_dicts

	def _compute_skins_and_payouts(self, pars, players_df):
		included = players_df[players_df.get("Included") == True].reset_index(drop=True)
		skins_awarded = {f"H{i+1}": [] for i in range(HOLES)}
		carryover_on = self.carryover_var.get()

		# parse per-skin / total purse inputs
		tp = self.total_purse_var.get().strip()
		try:
			per_skin_input = float(self.per_skin_var.get() or 1)
		except Exception:
			per_skin_input = 1.0
		total_purse = None
		if tp != "":
			try:
				total_purse = float(tp)
			except Exception:
				total_purse = None

		carryover_units = 0
		hole_results = []
		for i in range(HOLES):
			hole = f"H{i+1}"
			par_for_hole = pars[i]
			col = included[hole].replace("", pd.NA).dropna()
			if col.empty:
				hole_results.append({
					"hole": hole,
					"lowest": None,
					"tied": [],
					"sole_winner": None,
					"units_paid": 0,
					"carryover_before": carryover_units,
					"reason": REASON_NO_SCORES,
					"reason_text": "No scores"
				})
				continue
			col_int = col.astype(int)
			minv = col_int.min()
			tied_idx = col_int[col_int == minv].index.tolist()
			tied_names = list(included.loc[tied_idx, "Name"])

			# only par or less can win a skin
			if minv > par_for_hole:
				carryover_units += 1
				hole_results.append({
					"hole": hole,
					"lowest": minv,
					"tied": tied_names,
					"sole_winner": None,
					"units_paid": 0,
					"carryover_before": carryover_units - 1,
					"reason": REASON_CARRY,
					"reason_text": "All worse than par - carryover"
				})
				continue

			if len(tied_names) == 1:
				# sole winner: base unit + any carryover
				# optionally apply birdie/eagle bonus skins when enabled
				extra = 0
				if self.bonus_enabled_var.get():
					# minv is the winning score for the hole
					if minv == par_for_hole - 1:
						extra = 1
					elif minv <= par_for_hole - 2:
						extra = 5
				units = 1 + carryover_units + extra
				skins_awarded[hole].append((tied_names[0], units))
				hole_results.append({
					"hole": hole,
					"lowest": minv,
					"tied": tied_names,
					"sole_winner": tied_names[0],
					"units_paid": units,
					"carryover_before": carryover_units,
					"reason": REASON_SOLE,
					"reason_text": "Sole winner <= par"
				})
				carryover_units = 0
			else:
				# Tie case: decide whether to carry or split.
				# Rules:
				# - If tied score is worse than birdie (minv > par-1) -> carry
				# - If tied score is birdie or better (minv <= par-1):
				#     * If more than 3 players have birdie-or-better -> carry
				#     * Otherwise split the single unit among tied players
				if minv > par_for_hole - 1:
					# all worse than birdie -> carry
					carryover_units += 1
					hole_results.append({
						"hole": hole,
						"lowest": minv,
						"tied": tied_names,
						"sole_winner": None,
						"units_paid": 0,
						"carryover_before": carryover_units - 1,
						"reason": REASON_CARRY,
						"reason_text": "Tie worse than birdie -> carry"
					})
				else:
					# tied at birdie or better
					n_tied = len(tied_names)
					if n_tied > 3:
						# too many birdies/eagles -> carry
						carryover_units += 1
						hole_results.append({
							"hole": hole,
							"lowest": minv,
							"tied": tied_names,
							"sole_winner": None,
							"units_paid": 0,
							"carryover_before": carryover_units - 1,
							"reason": REASON_CARRY,
							"reason_text": "More than 3 birdies/eagles -> carry"
						})
					else:
						# split the single unit among tied players (no bonuses applied on ties)
						hole_results.append({
							"hole": hole,
							"lowest": minv,
							"tied": tied_names,
							"sole_winner": None,
							"units_paid": 1,
							"carryover_before": carryover_units,
							"reason": REASON_SPLIT,
							"reason_text": "Tie at birdie/eagle -> split"
						})
						# a skin (even split) was awarded, reset carryover counter
						carryover_units = 0

		payout_map_units = {name: 0.0 for name in included["Name"].tolist()}
		# First compute units-per-player only. Amounts depend on whether total_purse mode is used.
		for rec in hole_results:
			if rec.get("sole_winner"):
				winner = rec["sole_winner"]
				units = rec["units_paid"]
				payout_map_units[winner] += units
			else:
				# split is canonical reason
				if rec.get("reason") == REASON_SPLIT and rec.get("tied"):
					tied = rec["tied"]
					share_units = 1.0 / len(tied)
					for name in tied:
						payout_map_units[name] += share_units

		carryover_remaining = 0
		if carryover_on:
			carry = 0
			for r in hole_results:
				if r.get("sole_winner"):
					carry = 0
				else:
					# canonical carry reason
					if r.get("reason") == REASON_CARRY:
						carry += 1
			carryover_remaining = carry

		# determine per-unit payout amount
		# EXCLUSIVE MODE: if a valid total_purse is provided, use it ONLY (total_purse mode).
		# Otherwise use per-skin input. Do NOT fall back from total_purse to per-skin.
		# only use total_purse mode when total_purse is a positive number (>0)
		if total_purse is not None and total_purse > 0:
			total_units = sum(payout_map_units.values())
			if total_units > 0:
				per_unit = total_purse / total_units
			else:
				# no units awarded -> per_unit is zero (no payout). Avoid dividing by zero.
				per_unit = 0.0
		else:
			per_unit = per_skin_input

		# compute amounts per player based on units and per_unit value
		payout_map_amount = {name: round(payout_map_units[name] * per_unit, 2) for name in payout_map_units}

		return {
			"per_skin": per_unit,
			"payout_map_units": payout_map_units,
			"payout_map_amount": payout_map_amount,
			"skins_awarded": skins_awarded,
			"hole_results": hole_results,
			"carryover_remaining": carryover_remaining
		}

	def export_to_excel(self):
		pars, players = self.collect_data()
		if len(players) < 2:
			messagebox.showwarning("Not enough players", "Enter at least 2 players with names")
			return
		df = pd.DataFrame(players)
		results = self._compute_skins_and_payouts(pars, df)
		per_skin = results["per_skin"]
		payout_units = results["payout_map_units"]
		payout_amounts = results["payout_map_amount"]
		hole_results = results["hole_results"]
		carryover_remaining = results["carryover_remaining"]

		wb = Workbook()
		ws = wb.active
		ws.title = "Big Boy Skins Report"

		ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7 + HOLES)
		hcell = ws.cell(row=1, column=1)
		# include course name and date in the main header when available
		_course = (self.course_var.get().strip() if hasattr(self, 'course_var') else "")
		# determine date string (prefer DateEntry object if available)
		try:
			_date_str = self.date_entry.get_date().strftime("%Y-%m-%d")
		except Exception:
			_date_str = (self.date_var.get().strip() if hasattr(self, 'date_var') else "")
		# build header
		if _course and _date_str:
			hcell.value = f"Big Boy Skins at {_course} — {_date_str}"
		elif _course:
			hcell.value = f"Big Boy Skins at {_course}"
		elif _date_str:
			hcell.value = f"Big Boy Skins — {_date_str}"
		else:
			hcell.value = "Big Boy Skins"
		hcell.font = Font(size=20, bold=True)
		hcell.alignment = Alignment(horizontal="center")

		ws.cell(row=2, column=1, value="Course:")
		ws.cell(row=2, column=2, value=self.course_var.get().strip())
		ws.cell(row=3, column=1, value="Date:")
		try:
			ws.cell(row=3, column=2, value=self.date_entry.get_date().strftime("%Y-%m-%d"))
		except:
			ws.cell(row=3, column=2, value=self.date_var.get().strip())
		ws.cell(row=2, column=4, value="Per-skin $")
		ws.cell(row=2, column=5, value=float(per_skin))
		tp = self.total_purse_var.get().strip()
		if tp != "":
			try:
				ws.cell(row=3, column=4, value="Total Purse $")
				ws.cell(row=3, column=5, value=float(tp))
			except:
				pass
		ws.cell(row=4, column=4, value="Carryover Enabled")
		ws.cell(row=4, column=5, value=str(self.carryover_var.get()))

		start_row = 6
		headers = ["Name", "HCP", "Included"] + [f"H{i+1}" for i in range(HOLES)] + ["Front9", "Back9", "Units", "Amount$"]
		for c, h in enumerate(headers, start=1):
			cell = ws.cell(row=start_row, column=c, value=h)
			cell.font = Font(bold=True)
			# center-align score header cells (H1..H18) and front/back total headers
			if c >= 4 and c <= 4 + HOLES:
				cell.alignment = Alignment(horizontal="center")

		# Write Par and Stroke Index rows immediately below headers so the export contains
		# per-hole par and stroke index information aligned with H1..H18 columns.
		par_row = start_row + 1
		si_row = start_row + 2
		# style for par / stroke index rows to improve readability
		par_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
		si_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
		ws.cell(row=par_row, column=1, value="Par").font = Font(bold=True)
		ws.cell(row=par_row, column=1).fill = par_fill
		ws.cell(row=si_row, column=1, value="Stroke Index").font = Font(bold=True)
		ws.cell(row=si_row, column=1).fill = si_fill
		# Place pars and stroke indices under the H1..H18 columns (H1 column = 4)
		for i in range(HOLES):
			# pars were collected earlier
			try:
				pval = int(pars[i])
			except Exception:
				pval = pars[i] if pars[i] != None else ""
			cell_par = ws.cell(row=par_row, column=4 + i, value=pval)
			cell_par.alignment = Alignment(horizontal="center")
			cell_par.fill = par_fill
			# stroke index vars are stored as strings in self.stroke_index_vars
			try:
				siv = self.stroke_index_vars[i].get()
			except Exception:
				siv = ""
			cell_si = ws.cell(row=si_row, column=4 + i, value=siv)
			cell_si.alignment = Alignment(horizontal="center")
			cell_si.fill = si_fill

		participants = df[df.get("Included") == True]
		non_participants = df[df.get("Included") == False]
		write_order = pd.concat([participants, non_participants], ignore_index=True)
		# Data rows start after headers + Par + Stroke Index rows
		for r, (_, row) in enumerate(write_order.iterrows(), start=start_row+3):
			name = row.get("Name", "")
			included_flag = row.get("Included", False)
			ws.cell(row=r, column=1, value=name)
			# handicap if present
			hcp = row.get("Handicap", "")
			ws.cell(row=r, column=2, value=hcp if hcp != "" else None)
			ws.cell(row=r, column=3, value=included_flag)
			for i in range(HOLES):
				v = row.get(f"H{i+1}", "")
				cell = ws.cell(row=r, column=4 + i, value=v if v != "" else None)
				# center align per-hole score cells
				cell.alignment = Alignment(horizontal="center")
			# front/back totals: center-align as well
			cell_f = ws.cell(row=r, column=4 + HOLES, value=row.get("Front9", ""))
			cell_f.alignment = Alignment(horizontal="center")
			cell_b = ws.cell(row=r, column=5 + HOLES, value=row.get("Back9", ""))
			cell_b.alignment = Alignment(horizontal="center")
			units = payout_units.get(name, 0.0)
			amount = payout_amounts.get(name, 0.0)
			ws.cell(row=r, column=6 + HOLES, value=units if units != 0 else None)
			ws.cell(row=r, column=7 + HOLES, value=round(amount, 2) if amount != 0 else None)

		# highlight birdies/eagles
		birdie_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
		eagle_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
		for rr in range(start_row+3, start_row+3 + len(write_order)):
			for hi in range(HOLES):
				cell = ws.cell(row=rr, column=4 + hi)
				if cell.value is None:
					continue
				try:
					sv = int(cell.value)
				except:
					continue
				par_for_hole = pars[hi]
				if sv == par_for_hole - 1:
					cell.fill = birdie_fill
				elif sv <= par_for_hole - 2:
					cell.fill = eagle_fill

		# summary
		# move summary a few rows lower so there's spacing after the data table
		summary_row = start_row + 3 + len(write_order) + 4
		ws.cell(row=summary_row, column=1, value="Skins Summary").font = Font(bold=True)
		summary_row += 1
		ws.cell(row=summary_row, column=1, value="Hole").font = Font(bold=True)
		ws.cell(row=summary_row, column=2, value="Result").font = Font(bold=True)
		ws.cell(row=summary_row, column=3, value="Units Awarded").font = Font(bold=True)
		summary_row += 1
		for hr in results["hole_results"]:
			hole = hr["hole"]
			if hr.get("sole_winner"):
				text = hr["sole_winner"]
				units = hr["units_paid"]
			elif hr.get("tied"):
				r = hr.get("reason")
				if r == REASON_CARRY:
					# carry cases (could be many players or all worse than par)
					text = hr.get("reason_text", f"Tie ({', '.join(hr['tied'])}) - carryover")
					units = 0
				elif r == REASON_SPLIT:
					text = f"Tie ({', '.join(hr['tied'])}) - split"
					units = hr["units_paid"]
				elif r == REASON_NO_SCORES:
					text = hr.get("reason_text", "No scores")
					units = 0
				else:
					text = hr.get("reason_text", "No scores")
					units = 0
			else:
				text = hr.get("reason_text", "No scores")
				units = 0
			ws.cell(row=summary_row, column=1, value=hole)
			ws.cell(row=summary_row, column=2, value=text)
			ws.cell(row=summary_row, column=3, value=units)
			summary_row += 1

		# payouts
		summary_row += 1
		ws.cell(row=summary_row, column=1, value="Payouts").font = Font(bold=True)
		summary_row += 1
		ws.cell(row=summary_row, column=1, value="Name").font = Font(bold=True)
		ws.cell(row=summary_row, column=2, value="Units").font = Font(bold=True)
		ws.cell(row=summary_row, column=3, value="Amount$").font = Font(bold=True)
		summary_row += 1
		for name in participants["Name"].tolist():
			units = payout_units.get(name, 0.0)
			amount = payout_amounts.get(name, 0.0)
			ws.cell(row=summary_row, column=1, value=name)
			ws.cell(row=summary_row, column=2, value=round(units, 3) if units != 0 else None)
			ws.cell(row=summary_row, column=3, value=round(amount, 2) if amount != 0 else None)
			summary_row += 1
		if carryover_remaining:
			ws.cell(row=summary_row, column=1, value=f"Carryover units remaining after 18: {carryover_remaining}").font = Font(bold=True)
			summary_row += 1

		for c in range(1, 8 + HOLES):
			ws.column_dimensions[get_column_letter(c)].width = 14

		# create an Export Summary sheet with key settings so reports are self-describing
		try:
			summary = wb.create_sheet(title="Export Summary")
			summary.cell(row=1, column=1, value="Setting").font = Font(bold=True)
			summary.cell(row=1, column=2, value="Value").font = Font(bold=True)
			r = 2
			def srow(k, v):
				nonlocal r
				summary.cell(row=r, column=1, value=k)
				summary.cell(row=r, column=2, value=v)
				r += 1
			# basic metadata
			srow("Course", self.course_var.get().strip())
			try:
				srow("Date", self.date_entry.get_date().strftime("%Y-%m-%d"))
			except Exception:
				srow("Date", self.date_var.get().strip())
			srow("Per-skin $", float(per_skin))
			if tp != "":
				try:
					srow("Total Purse $", float(tp))
				except Exception:
					srow("Total Purse $", tp)
		# (No bonus points configured) - exporter will not include birdie/eagle bonus settings
				srow("Carryover Enabled", str(self.carryover_var.get()))
				srow("Use Net Scores", str(self.use_net_scores.get()))
				srow("Bonuses Enabled", str(self.bonus_enabled_var.get()))
				srow("Split Ties", str(self.split_ties.get()))

			# Add per-player summary: total units, bonus units, birdies, eagles
			try:
				# gather player names (Included True first)
				participants_df = df[df.get("Included") == True]
				player_names = participants_df.get("Name").tolist()
				# initialize counters
				units_map = results.get("payout_map_units", {})
				birdie_map = {n: 0 for n in player_names}
				eagle_map = {n: 0 for n in player_names}
				# count birdies/eagles from participant scores and pars
				for _, prow in participants_df.iterrows():
					name = prow.get("Name")
					for i in range(HOLES):
						try:
							val = prow.get(f"H{i+1}")
							if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
								continue
							score = int(val)
						except Exception:
							continue
						parv = int(pars[i]) if i < len(pars) else 4
						if score == parv - 1:
							birdie_map[name] += 1
						elif score <= parv - 2:
							eagle_map[name] += 1
				# write table header
				r += 1
				summary.cell(row=r, column=1, value="Player").font = Font(bold=True)
				summary.cell(row=r, column=2, value="Total Units").font = Font(bold=True)
				summary.cell(row=r, column=3, value="Birdies").font = Font(bold=True)
				summary.cell(row=r, column=4, value="Eagles").font = Font(bold=True)
				summary.cell(row=r, column=5, value="Amount$").font = Font(bold=True)
				r += 1
				# write per-player rows and accumulate total paid
				total_paid = 0.0
				for name in player_names:
					u = units_map.get(name, 0.0)
					amt = payout_amounts.get(name, 0.0)
					summary.cell(row=r, column=1, value=name)
					summary.cell(row=r, column=2, value=float(u) if u != 0 else None)
					summary.cell(row=r, column=3, value=birdie_map.get(name, 0) if birdie_map.get(name, 0) != 0 else None)
					summary.cell(row=r, column=4, value=eagle_map.get(name, 0) if eagle_map.get(name, 0) != 0 else None)
					cell_amt = summary.cell(row=r, column=5, value=round(amt or 0.0, 2))
					# always format amount as currency, even when zero
					cell_amt.number_format = '$#,##0.00'
					total_paid += float(amt or 0.0)
					r += 1
				# write total paid row
					if total_paid:
						summary.cell(row=r, column=1, value="Total Paid").font = Font(bold=True)
						cell_total = summary.cell(row=r, column=5, value=round(total_paid, 2))
						cell_total.number_format = '$#,##0.00'
						cell_total.font = Font(bold=True)
			except Exception:
				# ignore per-player summary failures
				pass
		except Exception:
			# non-fatal: if creating summary fails, continue with export
			pass

		# sanitize course name for filename
		safe_course = ''.join(c for c in (self.course_var.get().strip() or 'course') if c.isalnum() or c in (' ', '_', '-')).replace(' ', '_')
		default_name = f"BigBoySkins_{safe_course}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
		path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name,
						filetypes=[("Excel files", "*.xlsx")])
		if not path:
			return
		wb.save(path)
		messagebox.showinfo("Exported", f"Report exported to {path}")

	def import_from_excel(self):
		path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
		if not path:
			return
		try:
			wb = load_workbook(path, data_only=True)
			ws = wb.active
			# If the workbook contains an Export Summary sheet, read metadata from it
			try:
				summary_sheet = None
				for name in wb.sheetnames:
					if name.strip().lower() == "export summary" or name.strip().lower() == "export_summary":
						summary_sheet = wb[name]
						break
				if summary_sheet is not None:
					def _to_bool(v):
						if isinstance(v, bool):
							return v
						if v is None:
							return False
						s = str(v).strip().lower()
						return s in ("true", "1", "yes")
					for row in summary_sheet.iter_rows(values_only=True):
						if not row or row[0] is None:
							continue
						key = str(row[0]).strip().lower()
						val = row[1] if len(row) > 1 else None
						try:
							if "per-skin" in key:
								self.per_skin_var.set(str(val))
							elif "total purse" in key:
								self.total_purse_var.set(str(val))
							# no birdie/eagle bonus fields expected anymore
							elif "carryover" in key:
								self.carryover_var.set(_to_bool(val))
							elif "use net" in key:
								self.use_net_scores.set(_to_bool(val))
							elif "split ties" in key:
								self.split_ties.set(_to_bool(val))
							elif "course" == key:
								self.course_var.set(str(val))
							elif "date" == key and val is not None:
								try:
									# try to parse date objects or yyyy-mm-dd strings
									if hasattr(val, 'strftime'):
										self.date_entry.set_date(val)
									else:
										self.date_entry.set_date(datetime.strptime(str(val), "%Y-%m-%d").date())
								except Exception:
									# fall back to simple string set
									self.date_var.set(str(val))
						except Exception:
							# ignore malformed summary rows
							pass
			except Exception:
				# non-fatal: ignore summary read failures
						pass

			# find header row with "Name" in first column
			hr = None
			for r in range(1, 50):
				val = ws.cell(row=r, column=1).value
				if val == "Name":
					hr = r
					break
			if hr is None:
				messagebox.showerror("Import failed", "Could not find header row in spreadsheet")
				return

			rows = []
			r = hr + 1
			# Detect optional Par and Stroke Index rows written by export_to_excel
			first_label = ws.cell(row=r, column=1).value
			if isinstance(first_label, str) and first_label.strip().lower() == "par":
				# read pars from columns H1..H18 (starting at col 4)
				for i in range(HOLES):
					try:
						val = ws.cell(row=r, column=4 + i).value
						if val is None:
							self.par_vars[i].set("4")
						else:
							self.par_vars[i].set(str(int(val)) if isinstance(val, (int, float)) else str(val))
					except Exception:
						self.par_vars[i].set("4")
				r += 1

			second_label = ws.cell(row=r, column=1).value
			if isinstance(second_label, str) and second_label.strip().lower() in ("stroke index", "stroke_index", "si"):
				# read stroke indices from columns H1..H18
				for i in range(HOLES):
					try:
						val = ws.cell(row=r, column=4 + i).value
						if val is None:
							self.stroke_index_vars[i].set(str(i+1))
						else:
							self.stroke_index_vars[i].set(str(val))
					except Exception:
						self.stroke_index_vars[i].set(str(i+1))
				r += 1

			# Now r points to the first player row
			def _is_bool_like(val):
				if isinstance(val, bool):
					return True
				if isinstance(val, (int, float)) and val in (0, 1):
					return True
				try:
					s = str(val).strip().lower()
					if s in ("true", "false", "1", "0", "yes", "no"):
						return True
				except Exception:
					pass
				return False

			def _row_has_scores(ws, row):
				# Check H1..H18 columns for any non-empty or numeric entry
				for i in range(HOLES):
					v = ws.cell(row=row, column=4 + i).value
					if v is None:
						continue
					# consider numeric or any non-empty string a score
					if isinstance(v, (int, float)):
						return True
					if str(v).strip() != "":
						return True
				return False

			while True:
				name = ws.cell(row=r, column=1).value
				if name is None:
					break
				# stop if we hit clear summary headers
				if isinstance(name, str) and name.strip().lower() in ("skins summary", "payouts"):
					break
				included_cell = ws.cell(row=r, column=3).value
				hcp_cell = ws.cell(row=r, column=2).value
				# Decide if this row appears to be a player row.
				# Accept row if Included cell is boolean-like OR H1..H18 has any score OR handicap column is present/non-empty
				if not (_is_bool_like(included_cell) or _row_has_scores(ws, r) or (hcp_cell is not None and str(hcp_cell).strip() != "")):
					# not a player row -> stop importing
					break
				included = included_cell
				entry = {"Name": str(name), "Included": bool(included)}
				# try to read handicap
				hcp = ws.cell(row=r, column=2).value
				entry["Handicap"] = hcp if hcp is not None else "0"
				for i in range(HOLES):
					entry[f"H{i+1}"] = ws.cell(row=r, column=4 + i).value
				entry["Front9"] = ws.cell(row=r, column=4 + HOLES).value
				entry["Back9"] = ws.cell(row=r, column=5 + HOLES).value
				rows.append(entry)
				r += 1

			for p in self.players:
				p.destroy()
			self.players = []
			for i, row in enumerate(rows):
				pr = PlayerRow(self.player_frame, i, self)
				pr.load_from_dict(row)
				self.players.append(pr)
			while len(self.players) < 2:
				self.add_player()
			# import metadata
			try:
				course = ws.cell(row=2, column=2).value
				date_val = ws.cell(row=3, column=2).value
				if course:
					self.course_var.set(str(course))
				if isinstance(date_val, datetime):
					try:
						self.date_entry.set_date(date_val.date())
					except:
						pass
				else:
					try:
						self.date_entry.set_date(datetime.strptime(str(date_val), "%Y-%m-%d").date())
					except:
						pass
			except:
				pass

			messagebox.showinfo("Imported", "Contest imported successfully")
		except Exception as e:
			messagebox.showerror("Import error", f"Failed to import: {e}")


def main():
    root = tk.Tk()
    app = BigBoySkinsApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
