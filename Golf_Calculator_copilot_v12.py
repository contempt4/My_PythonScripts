import ctypes
import sys

# Fix blurry text on high-DPI displays (Windows)
# Created this fromV11 because of conflicts in onedrive
try:
    if sys.platform == 'win32':
        # Tell Windows this app is DPI-aware
        ctypes.windll.shcore.SetProcessDpiAwareness(1)  # 1 = System DPI aware
except Exception:
    pass

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


VERSION = "V1.0"
APP_ICON_FILENAME = "Golf_Icon.png"


def get_app_icon_path():
    """Return icon path for script mode and PyInstaller-frozen mode."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_dir = sys._MEIPASS
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, APP_ICON_FILENAME)

MAX_PLAYERS = 40
MAX_HOLE_SCORE = 9
HOLES = 18

# canonical reason codes for hole results
REASON_NO_SCORES = "NO_SCORES"
REASON_SOLE = "SOLE"
REASON_CARRY = "CARRY"
REASON_SPLIT = "SPLIT"


class PlayerRow:
    def __init__(self, parent, idx, app):
        # place widgets directly into the parent grid so their columns align with header labels
        self.parent = parent
        self.app = app
        # players start after Par (row0), Stroke Index (row1), Header (row2)
        self.row = idx + 3
        self.name_var = tk.StringVar()
        self.handicap_var = tk.StringVar(value="0")
        self.include_var = tk.BooleanVar(value=True)
        self.score_vars = [tk.StringVar(value="") for _ in range(HOLES)]
        self.score_entries = []
        self.score_entry_defaults = []

        self.name_entry = ttk.Entry(self.parent, textvariable=self.name_var, width=20)
        self.name_entry.grid(row=self.row, column=0, padx=2, pady=2)

        self.handicap_entry = ttk.Entry(self.parent, textvariable=self.handicap_var, width=5)
        self.handicap_entry.grid(row=self.row, column=1, padx=2, pady=2)

        self.include_cb = ttk.Checkbutton(self.parent, variable=self.include_var)
        self.include_cb.grid(row=self.row, column=2, padx=2, pady=2)

        for h in range(HOLES):
            e = tk.Entry(self.parent, textvariable=self.score_vars[h], width=3)
            col = 3 + h if h < 9 else 4 + h
            e.grid(row=self.row, column=col, padx=1, pady=2)
            try:
                self.score_entry_defaults.append(e.cget('bg'))
            except Exception:
                self.score_entry_defaults.append(None)
            self.score_entries.append(e)
            e.bind("<FocusOut>", lambda ev, sv=self.score_vars[h]: self._validate_score_var(sv))

        self.front9_lbl = ttk.Label(self.parent, text="0", width=5)
        self.front9_lbl.grid(row=self.row, column=12, padx=4)
        self.back9_lbl = ttk.Label(self.parent, text="0", width=5)
        self.back9_lbl.grid(row=self.row, column=4 + HOLES, padx=4)

        for sv in self.score_vars:
            sv.trace_add("write", lambda *a: self.update_totals())

    def _validate_score_var(self, var):
        v = var.get().strip()
        if v == "":
            return
        try:
            n = int(v)
            if n < 0 or n > MAX_HOLE_SCORE:
                messagebox.showwarning("Invalid score", f"Score must be between 0 and {MAX_HOLE_SCORE}")
                var.set("")
        except ValueError:
            messagebox.showwarning("Invalid score", "Score must be an integer")
            var.set("")

    def update_totals(self):
        front = 0
        back = 0
        for i in range(9):
            v = self.score_vars[i].get().strip()
            if v.isdigit():
                front += int(v)
        for i in range(9, 18):
            v = self.score_vars[i].get().strip()
            if v.isdigit():
                back += int(v)
        self.front9_lbl.config(text=str(front))
        self.back9_lbl.config(text=str(back))

        # Highlight birdies/eagles
        for i in range(HOLES):
            try:
                ent = self.score_entries[i]
            except Exception:
                continue
            val = self.score_vars[i].get().strip()
            if val == "":
                try:
                    default = self.score_entry_defaults[i]
                    ent.config(bg=default if default is not None else 'white')
                except Exception:
                    pass
                continue
            try:
                score = int(val)
            except Exception:
                continue
            try:
                par_v = int(self.app.par_vars[i].get())
            except Exception:
                par_v = 4
            if score == par_v - 1:
                try:
                    ent.config(bg="#FFF59D")
                except Exception:
                    pass
            elif score <= par_v - 2:
                try:
                    ent.config(bg="#C8E6C9")
                except Exception:
                    pass
            else:
                try:
                    default = self.score_entry_defaults[i]
                    ent.config(bg=default if default is not None else 'white')
                except Exception:
                    pass

    def destroy(self):
        try:
            self.name_entry.destroy()
        except:
            pass
        try:
            self.handicap_entry.destroy()
        except:
            pass
        try:
            self.include_cb.destroy()
        except:
            pass
        for h in range(HOLES):
            try:
                col = 3 + h if h < 9 else 4 + h
                w = self.parent.grid_slaves(row=self.row, column=col)
                if w:
                    w[0].destroy()
            except:
                pass
        try:
            self.front9_lbl.destroy()
        except:
            pass
        try:
            self.back9_lbl.destroy()
        except:
            pass

    def to_dict(self):
        d = {
            "Name": self.name_var.get().strip(),
            "Handicap": self.handicap_var.get().strip(),
            "Included": self.include_var.get()
        }
        for i in range(HOLES):
            d[f"H{i+1}"] = self.score_vars[i].get().strip()
        d["Front9"] = self.front9_lbl.cget("text")
        d["Back9"] = self.back9_lbl.cget("text")
        return d

    def load_from_dict(self, d):
        self.name_var.set(d.get("Name", ""))
        self.handicap_var.set(str(d.get("Handicap", "0")))
        self.include_var.set(bool(d.get("Included", True)))
        for i in range(HOLES):
            key = f"H{i+1}"
            val = d.get(key, "")
            if pd.isna(val):
                self.score_vars[i].set("")
            else:
                self.score_vars[i].set(str(val))
        self.update_totals()


class BigBoySkinsApp:
    def __init__(self, root):
        self.root = root
        self.root.title(f"Big Boys Skins Manager  {VERSION}")
        self.players = []
        self.use_net_scores = tk.BooleanVar(value=False)
        self.split_ties = tk.BooleanVar(value=False)
        self.course_name_var = tk.StringVar()
        self.course_var = self.course_name_var
        self.date_var = tk.StringVar()
        self.per_skin_var = tk.StringVar(value="1")
        self.total_purse_var = tk.StringVar(value="")
        self.carryover_var = tk.BooleanVar(value=True)
        self.bonus_enabled_var = tk.BooleanVar(value=True)
        self.par_vars = [tk.StringVar(value="4") for _ in range(HOLES)]
        self.stroke_index_vars = [tk.StringVar(value=str(i+1)) for i in range(HOLES)]

        self.build_gui()

   # ...existing code...
    def build_gui(self):
        # Large in-app title (can't change OS title bar font from Tkinter)
        title_lbl = ttk.Label(self.root, text=self.root.title() or f"Big Boys Skins Manager  {VERSION}",
                      font=("Segoe UI", 14, "bold"))
        title_lbl.grid(row=0, column=0, sticky="w", padx=12, pady=(8, 4))

        header = ttk.Frame(self.root)
        header.grid(row=1, column=0, sticky="w", padx=10, pady=5)
        # add internal padding to the header and increase gaps between widgets
        header = ttk.Frame(self.root, padding=(12, 8))
        header.grid(row=1, column=0, sticky="w", padx=12, pady=8)
        style = ttk.Style()
        style.configure('.', font=('Segoe UI', 11))
        self.root.columnconfigure(0, weight=1)
        header.columnconfigure(1, weight=2)
        header.columnconfigure(3, weight=1)
        header.columnconfigure(5, weight=1)
        header.columnconfigure(7, weight=1)
        # make the player list row (now row 2) expandable
        self.root.rowconfigure(2, weight=1)
        
       
        #ttk.Label(header, text="Course Name:").grid(row=0, column=0)
        ttk.Entry(header, textvariable=self.course_name_var, width=30).grid(row=0, column=1, sticky="ew")
        ttk.Label(header, text="Course Name:").grid(row=0, column=0, padx=(0,8), pady=4)
        ttk.Entry(header, textvariable=self.course_name_var, width=36).grid(row=0, column=1, sticky="ew", padx=(0,12), pady=4)

        #ttk.Label(header, text="Date:").grid(row=0, column=2)
        #self.date_entry = DateEntry(header, textvariable=self.date_var, width=12)
        #self.date_entry.grid(row=0, column=3, sticky="ew")
        ttk.Label(header, text="Date:").grid(row=0, column=2, padx=(6,4), pady=4)
        self.date_entry = DateEntry(header, textvariable=self.date_var, width=14)
        self.date_entry.grid(row=0, column=3, sticky="ew", padx=(0,12), pady=4)

        #tk.Label(header, text="Per Skin $").grid(row=0, column=4)
        ttk.Entry(header, textvariable=self.per_skin_var, width=6).grid(row=0, column=5, sticky="ew")
        ttk.Label(header, text="Total Purse $").grid(row=0, column=6)
        ttk.Entry(header, textvariable=self.total_purse_var, width=8).grid(row=0, column=7, sticky="ew")
        #ttk.Checkbutton(header, text='Carryover', variable=self.carryover_var).grid(row=0, column=8)
        ttk.Checkbutton(header, text='Apply Birdie/Eagle Bonuses', variable=self.bonus_enabled_var).grid(row=0, column=9)
        ttk.Label(header, text="Per Skin $").grid(row=0, column=4, padx=(6,4), pady=4)
        ttk.Entry(header, textvariable=self.per_skin_var, width=8).grid(row=0, column=5, sticky="ew", padx=(0,12), pady=4)
        #tk.Label(header, text="Total Purse $").grid(row=0, column=6, padx=(6,4), pady=4)
        ttk.Entry(header, textvariable=self.total_purse_var, width=10).grid(row=0, column=7, sticky="ew", padx=(0,12), pady=4)
        ttk.Checkbutton(header, text='Carryover', variable=self.carryover_var).grid(row=0, column=8, padx=(6,8), pady=4)
        ttk.Checkbutton(header, text='Apply Birdie/Eagle Bonuses', variable=self.bonus_enabled_var).grid(row=0, column=9, padx=(0,8), pady=4)

        #ttk.Checkbutton(header, text="Use Net Scores (based on handicap)", variable=self.use_net_scores).grid(row=1, column=0, columnspan=2, sticky="w")
        #ttk.Checkbutton(header, text="Split Ties (instead of carryover)", variable=self.split_ties).grid(row=1, column=2, columnspan=2, sticky="w")
        ttk.Checkbutton(header, text="Use Net Scores (based on handicap)", variable=self.use_net_scores).grid(row=1, column=0, columnspan=2, sticky="w", padx=(0,8), pady=6)
        ttk.Checkbutton(header, text="Split Ties (instead of carryover)", variable=self.split_ties).grid(row=1, column=2, columnspan=2, sticky="w", padx=(6,8), pady=6)
 # ...existing code...

        # Player list: make it scrollable. Container holds a Canvas and vertical Scrollbar.
        container = ttk.Frame(self.root)
        container.grid(row=2, column=0, sticky="nsew", padx=10, pady=5)
        container.columnconfigure(0, weight=1)
        # allow the canvas (row 0) to expand vertically when the window is resized
        container.rowconfigure(0, weight=1)

        self.player_canvas = tk.Canvas(container, highlightthickness=0)
        self.player_vscroll = ttk.Scrollbar(container, orient="vertical", command=self.player_canvas.yview)
        self.player_canvas.configure(yscrollcommand=self.player_vscroll.set)

        self.player_canvas.grid(row=0, column=0, sticky="nsew")
        self.player_vscroll.grid(row=0, column=1, sticky="ns")

        # inner frame where rows live
        self.player_inner = ttk.Frame(self.player_canvas)
        self.player_inner_id = self.player_canvas.create_window((0, 0), window=self.player_inner, anchor='nw')

        # column weights on inner frame (same as previous player_frame)
        self.player_inner.columnconfigure(0, weight=2)
        for ci in range(1, 8 + HOLES):
            self.player_inner.columnconfigure(ci, weight=1)

        # ensure canvas expands properly when container size changes
        def _on_inner_config(event):
            try:
                self.player_canvas.configure(scrollregion=self.player_canvas.bbox("all"))
            except Exception:
                pass
        self.player_inner.bind('<Configure>', _on_inner_config)

        def _on_canvas_config(event):
            # keep inner frame width in sync with canvas width
            try:
                cw = event.width
                self.player_canvas.itemconfigure(self.player_inner_id, width=cw)
            except Exception:
                pass
        self.player_canvas.bind('<Configure>', _on_canvas_config)

        # mousewheel scrolling (Windows behavior)
        def _on_mousewheel(event):
            # On Windows, event.delta is multiples of 120
            try:
                self.player_canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
            except Exception:
                pass
        self.player_canvas.bind_all('<MouseWheel>', _on_mousewheel)

        # Par row
        ttk.Label(self.player_inner, text="Par:").grid(row=0, column=0, sticky="w")
        for i in range(HOLES):
            col = 3 + i if i < 9 else 4 + i
            tk.Entry(self.player_inner, textvariable=self.par_vars[i], width=3).grid(row=0, column=col)

        # Stroke Index row
        ttk.Label(self.player_inner, text="Stroke Index:").grid(row=1, column=0, sticky="w")
        for i in range(HOLES):
            col = 3 + i if i < 9 else 4 + i
            tk.Entry(self.player_inner, textvariable=self.stroke_index_vars[i], width=3).grid(row=1, column=col)

        # Header row
        ttk.Label(self.player_inner, text="Name").grid(row=2, column=0)
        ttk.Label(self.player_inner, text="HCP").grid(row=2, column=1)
        ttk.Label(self.player_inner, text="In?").grid(row=2, column=2)
        for i in range(HOLES):
            col = 3 + i if i < 9 else 4 + i
            ttk.Label(self.player_inner, text=f"H{i+1}").grid(row=2, column=col)
        ttk.Label(self.player_inner, text="Front9").grid(row=2, column=12)
        ttk.Label(self.player_inner, text="Back9").grid(row=2, column=4 + HOLES)

        # Buttons
        btn_frame = ttk.Frame(self.root)
        btn_frame.grid(row=3, column=0, sticky="w", padx=10, pady=10)
        ttk.Button(btn_frame, text="Add Player", command=self.add_player).grid(row=0, column=0, padx=5)
        ttk.Button(btn_frame, text="Export to Excel", command=self.export_to_excel).grid(row=0, column=1, padx=5)
        ttk.Button(btn_frame, text="Import from Excel", command=self.import_from_excel).grid(row=0, column=2, padx=5)

        for _ in range(2):
            self.add_player()

    def _adjust_height(self):
        """Increase window height so player rows are visible. Caps at screen height minus a margin.

        Uses requested geometry to estimate needed height and resizes the root window while
        preserving current width.
        """
        try:
            self.root.update_idletasks()

            # If we have a scrollable player canvas, prefer the canvas content height
            content_h = None
            try:
                if hasattr(self, 'player_canvas') and self.player_canvas is not None:
                    # ensure scrollregion is up-to-date
                    try:
                        self.player_canvas.configure(scrollregion=self.player_canvas.bbox('all'))
                    except Exception:
                        pass
                    bbox = self.player_canvas.bbox('all')
                    if bbox:
                        # bbox = (x1, y1, x2, y2)
                        content_h = int(bbox[3] - bbox[1])
            except Exception:
                content_h = None

            if content_h and content_h > 0:
                # estimate header + controls area height as a constant buffer
                header_buffer = 420
                desired_h = content_h + header_buffer
            else:
                # fallback to requested height of all widgets
                req_h = self.root.winfo_reqheight()
                desired_h = req_h + 24

            screen_h = self.root.winfo_screenheight()
            max_h = max(200, screen_h - 100)
            new_h = min(desired_h, max_h)

            # preserve current width if available, otherwise default to 1200
            cur_w = self.root.winfo_width()
            if not cur_w or cur_w < 100:
                # try reading geometry string fallback
                try:
                    geom = self.root.winfo_geometry()
                    cur_w = int(geom.split('x')[0])
                except Exception:
                    cur_w = 2040

            self.root.geometry(f"{cur_w}x{int(new_h)}")
        except Exception:
            # non-fatal; don't block adding players if resizing fails
            pass

    def add_player(self):
        if len(self.players) >= MAX_PLAYERS:
            messagebox.showwarning("Limit reached", f"Maximum {MAX_PLAYERS} players allowed.")
            return
        idx = len(self.players)
        row = PlayerRow(self.player_inner, idx, self)
        self.players.append(row)
        # adjust window height so new row is visible (capped to screen size)
        try:
            self._adjust_height()
        except Exception:
            pass

    def collect_data(self):
        pars = []
        for v in self.par_vars:
            s = v.get().strip()
            pars.append(int(s) if s.isdigit() else 4)
        player_dicts = []
        for pr in self.players:
            d = pr.to_dict()
            if d.get("Name", "") == "":
                continue
            for i in range(HOLES):
                key = f"H{i+1}"
                val = d.get(key, "")
                if isinstance(val, str) and val.strip() == "":
                    d[key] = ""
                else:
                    try:
                        n = int(val)
                        if n < 0 or n > MAX_HOLE_SCORE:
                            d[key] = ""
                        else:
                            d[key] = n
                    except Exception:
                        d[key] = ""
            player_dicts.append(d)
        return pars, player_dicts

    def _format_bonus_summary(self, hole_result):
        bonus_map = hole_result.get("gross_bonus_map") or hole_result.get("bonus_map") or {}
        if not bonus_map:
            return ""

        parts = []
        for name, bonus in bonus_map.items():
            try:
                bonus_value = float(bonus)
            except Exception:
                continue
            if bonus_value == 0:
                continue
            bonus_text = str(int(bonus_value)) if bonus_value.is_integer() else f"{bonus_value:g}"
            parts.append(f"{name} +{bonus_text}")

        if not parts:
            return ""
        return f" (bonus units: {', '.join(parts)})"

    def _compute_skins_and_payouts(self, pars, players_df):
        included = players_df[players_df.get("Included") == True].reset_index(drop=True)
        # preserve gross scores (before handicap/net adjustments) so bonuses are based on gross
        gross_scores = included.copy(deep=True)

        if self.use_net_scores.get():
            try:
                hcp = pd.to_numeric(included["Handicap"], errors="coerce").fillna(0)
                stroke_index = [int(v.get()) for v in self.stroke_index_vars]
                strokes = []
                for idx, row in included.iterrows():
                    player_hcp = hcp.iloc[idx]
                    strokes_for_player = [1 if si <= player_hcp % 18 else 0 for si in stroke_index]
                    strokes_for_player = [s + (player_hcp // 18) for s in strokes_for_player]
                    strokes.append(strokes_for_player)
                for i, player_strokes in enumerate(strokes):
                    for h in range(HOLES):
                        col = f"H{h+1}"
                        if col in included.columns and included.at[i, col] not in ("", None):
                            try:
                                included.at[i, col] = int(included.at[i, col]) - player_strokes[h]
                            except Exception:
                                pass
            except Exception as e:
                print("Error applying handicaps:", e)

        skins_awarded = {f"H{i+1}": [] for i in range(HOLES)}
        carryover_on = self.carryover_var.get()

        tp = self.total_purse_var.get().strip()
        try:
            per_skin_input = float(self.per_skin_var.get() or 1)
        except Exception:
            per_skin_input = 1.0
        total_purse = None
        if tp != "":
            try:
                total_purse = float(tp)
            except Exception:
                total_purse = None

        carryover_units = 0
        hole_results = []
        for i in range(HOLES):
            hole = f"H{i+1}"
            par_for_hole = pars[i]
            col = included[hole].replace("", pd.NA).dropna()
            if col.empty:
                hole_results.append({
                    "hole": hole,
                    "lowest": None,
                    "tied": [],
                    "sole_winner": None,
                    "units_paid": 0,
                    "carryover_before": carryover_units,
                    "reason": REASON_NO_SCORES,
                    "reason_text": "No scores"
                })
                continue
            col_int = col.astype(int)
            minv = col_int.min()
            tied_idx = col_int[col_int == minv].index.tolist()
            tied_names = list(included.loc[tied_idx, "Name"])

            if minv > par_for_hole:
                carry_before = carryover_units
                if carryover_on:
                    carryover_units += 1
                    reason_text = "All worse than par - carryover"
                    reason = REASON_CARRY
                else:
                    reason_text = "All worse than par - carry disabled"
                    reason = REASON_NO_SCORES
                hole_results.append({
                    "hole": hole,
                    "lowest": minv,
                    "tied": tied_names,
                    "sole_winner": None,
                    "units_paid": 0,
                    "carryover_before": carry_before,
                    "reason": reason,
                    "reason_text": reason_text
                })
                continue

            if len(tied_names) == 1:
                extra = 0
                if self.bonus_enabled_var.get():
                    # bonus should be applied based on gross score (not net). If net scoring is enabled
                    # we stored the original gross values in `gross_scores` before adjustment above.
                    try:
                        if self.use_net_scores.get():
                            winner_idx = tied_idx[0]
                            gv = gross_scores.at[winner_idx, hole]
                            gross_val = int(gv) if (gv is not None and str(gv).strip() != "" and not pd.isna(gv)) else minv
                        else:
                            gross_val = minv
                    except Exception:
                        gross_val = minv

                    if gross_val == par_for_hole - 1:
                        extra = 1
                    elif gross_val <= par_for_hole - 2:
                        extra = 5
                units = 1 + carryover_units + extra
                skins_awarded[hole].append((tied_names[0], units))
                hole_result = {
                    "hole": hole,
                    "lowest": minv,
                    "tied": tied_names,
                    "sole_winner": tied_names[0],
                    "units_paid": units,
                    "carryover_before": carryover_units,
                    "reason": REASON_SOLE,
                    "reason_text": "Sole winner <= par"
                }
                if extra:
                    hole_result["gross_bonus_map"] = {tied_names[0]: extra}
                hole_results.append(hole_result)
                carryover_units = 0
            else:
                if minv > par_for_hole - 1:
                    carry_before = carryover_units
                    if carryover_on:
                        carryover_units += 1
                        reason_text = "Tie worse than birdie -> carry"
                        reason = REASON_CARRY
                    else:
                        reason_text = "Tie worse than birdie -> carry disabled"
                        reason = REASON_NO_SCORES
                    hole_results.append({
                        "hole": hole,
                        "lowest": minv,
                        "tied": tied_names,
                        "sole_winner": None,
                        "units_paid": 0,
                        "carryover_before": carry_before,
                        "reason": reason,
                        "reason_text": reason_text
                    })
                else:
                    n_tied = len(tied_names)
                    # If more than two players made birdie/eagle on the hole, treat as carry
                    # (no splits awarded when >2 birdies/eagles).
                    if n_tied > 2:
                        # For >2 birdies/eagles: each tied player retains their bonus (if any),
                        # and the hole becomes a carry (carry increases by 1 only).
                        bonus_map = {}
                        try:
                            for idx_t in tied_idx:
                                try:
                                    gv = gross_scores.at[idx_t, hole]
                                except Exception:
                                    gv = None
                                if gv is None or (isinstance(gv, float) and pd.isna(gv)) or str(gv).strip() == "":
                                    continue
                                try:
                                    gvi = int(gv)
                                except Exception:
                                    try:
                                        gvi = int(float(gv))
                                    except Exception:
                                        gvi = None
                                if gvi is None:
                                    continue
                                if gvi <= par_for_hole - 2:
                                    bonus_map[included.at[idx_t, "Name"]] = 5
                                elif gvi == par_for_hole - 1:
                                    bonus_map[included.at[idx_t, "Name"]] = 1
                        except Exception:
                            bonus_map = {}

                        carry_before = carryover_units
                        if carryover_on:
                            carryover_units += 1
                            if bonus_map:
                                reason_text = f"More than 2 birdies/eagles -> carry (bonuses retained)"
                            else:
                                reason_text = "More than 2 birdies/eagles -> carry"
                            reason = REASON_CARRY
                        else:
                            reason_text = "More than 2 birdies/eagles -> carry disabled"
                            reason = REASON_NO_SCORES
                        hole_results.append({
                            "hole": hole,
                            "lowest": minv,
                            "tied": tied_names,
                            "sole_winner": None,
                            "units_paid": 0,
                            "carryover_before": carry_before,
                            "reason": reason,
                            "reason_text": reason_text,
                            "bonus_map": bonus_map
                        })
                    else:
                        # Two-player tie at birdie/eagle: either split or carry depending on setting
                        if self.split_ties.get():
                            split_units = 1 + carryover_units
                            hole_results.append({
                                "hole": hole,
                                "lowest": minv,
                                "tied": tied_names,
                                "sole_winner": None,
                                "units_paid": split_units,
                                "carryover_before": carryover_units,
                                "reason": REASON_SPLIT,
                                "reason_text": "Tie at birdie/eagle -> split"
                            })
                            carryover_units = 0
                        else:
                            carry_before = carryover_units
                            if carryover_on:
                                carryover_units += 1
                                reason_text = "Split ties disabled -> carry"
                                reason = REASON_CARRY
                            else:
                                reason_text = "Split ties disabled -> carry disabled"
                                reason = REASON_NO_SCORES
                            hole_results.append({
                                "hole": hole,
                                "lowest": minv,
                                "tied": tied_names,
                                "sole_winner": None,
                                "units_paid": 0,
                                "carryover_before": carry_before,
                                "reason": reason,
                                "reason_text": reason_text
                            })

        # If carryover is disabled, ensure no hole is treated as a carry and
        # remove any retained bonus_map entries so no units are awarded.
        if not carryover_on:
            for rec in hole_results:
                if rec.get("reason") == REASON_CARRY:
                    rec["reason"] = REASON_NO_SCORES
                    # append a note if not already present
                    rt = rec.get("reason_text", "")
                    if "carry disabled" not in rt.lower():
                        rec["reason_text"] = (rt + " - carry disabled").strip()
                if "bonus_map" in rec:
                    # remove bonus_map when carry is disabled
                    rec.pop("bonus_map", None)

        payout_map_units = {name: 0.0 for name in included["Name"].tolist()}

        # Award unconditional gross-based bonuses (birdie=1, eagle=5) to all players
        # based on preserved `gross_scores`. Bonuses are awarded for ANY under-par score
        # regardless of whether a skin was won on that hole.
        if self.bonus_enabled_var.get():
            hole_to_rec = {rec.get("hole"): rec for rec in hole_results}
            for idx, prow in included.iterrows():
                pname = prow.get("Name")
                for h in range(HOLES):
                    hole = f"H{h+1}"
                    try:
                        gv = gross_scores.at[idx, hole]
                    except Exception:
                        gv = None
                    if gv is None or (isinstance(gv, float) and pd.isna(gv)) or str(gv).strip() == "":
                        continue
                    try:
                        gvi = int(gv)
                    except Exception:
                        try:
                            gvi = int(float(gv))
                        except Exception:
                            continue
                    par_for_hole = pars[h]
                    bonus = 0
                    if gvi <= par_for_hole - 2:
                        bonus = 5
                    elif gvi == par_for_hole - 1:
                        bonus = 1
                    if not bonus:
                        continue

                    rec = hole_to_rec.get(hole)
                    # Sole winners already have their birdie/eagle bonus folded into
                    # units_paid for that hole. Skip the later bonus pass to avoid
                    # counting the same bonus twice.
                    if rec is not None and rec.get("sole_winner") == pname:
                        continue

                    # Award bonus unconditionally for any under-par score
                    if pname in payout_map_units:
                        payout_map_units[pname] += float(bonus)
                    
                    # Record applied bonus in hole result for reporting
                    if rec is not None:
                        rec.setdefault("gross_bonus_map", {})[pname] = int(bonus)

        # Now allocate the standard hole payouts (sole winners and splits).
        for rec in hole_results:
            if rec.get("sole_winner"):
                winner = rec["sole_winner"]
                units = rec.get("units_paid", 0)
                if winner in payout_map_units:
                    payout_map_units[winner] += float(units)
            else:
                if rec.get("reason") == REASON_SPLIT and rec.get("tied"):
                    tied = rec["tied"]
                    split_units = float(rec.get("units_paid", 0))
                    share_units = split_units / len(tied) if len(tied) > 0 else 0
                    for name in tied:
                        if name in payout_map_units:
                            payout_map_units[name] += float(share_units)

        carryover_remaining = 0
        if carryover_on:
            carry = 0
            for r in hole_results:
                if r.get("sole_winner") or r.get("reason") == REASON_SPLIT:
                    carry = 0
                else:
                    if r.get("reason") == REASON_CARRY:
                        carry += 1
            carryover_remaining = carry

        if total_purse is not None and total_purse > 0:
            total_units = sum(payout_map_units.values())
            per_unit = (total_purse / total_units) if total_units > 0 else 0.0
        else:
            per_unit = per_skin_input

        payout_map_amount = {name: round(payout_map_units[name] * per_unit, 2) for name in payout_map_units}

        return {
            "per_skin": per_unit,
            "payout_map_units": payout_map_units,
            "payout_map_amount": payout_map_amount,
            "skins_awarded": skins_awarded,
            "hole_results": hole_results,
            "carryover_remaining": carryover_remaining
        }

   # ...existing code...
    def export_to_excel(self):
        pars, players = self.collect_data()
        if len(players) < 2:
            messagebox.showwarning("Not enough players", "Enter at least 2 players with names")
            return

        df = pd.DataFrame(players)
        results = self._compute_skins_and_payouts(pars, df)
        per_skin = results["per_skin"]
        payout_units = results["payout_map_units"]
        payout_amounts = results["payout_map_amount"]
        hole_results = results.get("hole_results", [])
        carryover_remaining = results.get("carryover_remaining", 0)

        wb = Workbook()
        ws = wb.active
        ws.title = "Big Boy Skins Report"
        # Leave column A blank to match user's example export (start content in column B)
        col_off = 1  # number of columns to shift right (1 => start at column B)

        ws.merge_cells(start_row=1, start_column=1 + col_off, end_row=1, end_column=(7 + HOLES) + col_off)
        hcell = ws.cell(row=1, column=1 + col_off)
        _course = (self.course_var.get().strip() if hasattr(self, 'course_var') else "")
        try:
            _date_str = self.date_entry.get_date().strftime("%Y-%m-%d")
        except Exception:
            _date_str = (self.date_var.get().strip() if hasattr(self, 'date_var') else "")
        if _course and _date_str:
            hcell.value = f"Big Boy Skins at {_course} — {_date_str}"
        elif _course:
            hcell.value = f"Big Boy Skins at {_course}"
        elif _date_str:
            hcell.value = f"Big Boy Skins — {_date_str}"
        else:
            hcell.value = "Big Boy Skins"
        hcell.font = Font(size=20, bold=True)
        hcell.alignment = Alignment(horizontal="center")

        # Metadata (shifted right by col_off)
        ws.cell(row=2, column=1 + col_off, value="Course:")
        ws.cell(row=2, column=2 + col_off, value=self.course_var.get().strip())
        ws.cell(row=3, column=1 + col_off, value="Date:")
        try:
            ws.cell(row=3, column=2 + col_off, value=self.date_entry.get_date().strftime("%Y-%m-%d"))
        except Exception:
            ws.cell(row=3, column=2 + col_off, value=self.date_var.get().strip())
        ws.cell(row=2, column=4 + col_off, value="Per-skin $")
        ws.cell(row=2, column=5 + col_off, value=float(per_skin))

        tp = self.total_purse_var.get().strip()
        if tp != "":
            try:
                ws.cell(row=3, column=4 + col_off, value="Total Purse $")
                ws.cell(row=3, column=5 + col_off, value=float(tp))
            except Exception:
                pass

        ws.cell(row=4, column=4 + col_off, value="Carryover Enabled")
        ws.cell(row=4, column=5 + col_off, value=str(self.carryover_var.get()))

        start_row = 6
        headers = ["Name", "HCP", "Included"] + [f"H{i+1}" for i in range(HOLES)] + ["Front9", "Back9", "Units", "Amount$"]

        # thin border for table cells
        thin = Side(border_style="thin", color="000000")
        bd = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=c + col_off, value=h)
            cell.font = Font(bold=True)
            cell.border = bd
            if 4 <= c <= 4 + HOLES:
                cell.alignment = Alignment(horizontal="center")

        par_row = start_row + 1
        si_row = start_row + 2
        par_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        si_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.cell(row=par_row, column=1 + col_off, value="Par").font = Font(bold=True)
        ws.cell(row=par_row, column=1 + col_off).fill = par_fill
        ws.cell(row=si_row, column=1 + col_off, value="Stroke Index").font = Font(bold=True)
        ws.cell(row=si_row, column=1 + col_off).fill = si_fill

        for i in range(HOLES):
            try:
                pval = int(pars[i])
            except Exception:
                pval = pars[i] if pars[i] is not None else ""
            cell_par = ws.cell(row=par_row, column=4 + i + col_off, value=pval)
            cell_par.alignment = Alignment(horizontal="center")
            cell_par.fill = par_fill

            try:
                siv = self.stroke_index_vars[i].get()
            except Exception:
                siv = ""
            cell_si = ws.cell(row=si_row, column=4 + i + col_off, value=siv)
            cell_si.alignment = Alignment(horizontal="center")
            cell_si.fill = si_fill

        participants = df[df.get("Included") == True]
        non_participants = df[df.get("Included") == False]
        write_order = pd.concat([participants, non_participants], ignore_index=True)

        # Data rows start after headers + Par + Stroke Index rows
        for r, (_, row) in enumerate(write_order.iterrows(), start=start_row + 3):
            name = row.get("Name", "")
            included_flag = row.get("Included", False)
            cell_name = ws.cell(row=r, column=1 + col_off, value=name)
            cell_name.border = bd
            hcp = row.get("Handicap", "")
            cell_hcp = ws.cell(row=r, column=2 + col_off, value=hcp if hcp != "" else None)
            cell_hcp.border = bd
            cell_in = ws.cell(row=r, column=3 + col_off, value=included_flag)
            cell_in.border = bd
            for i in range(HOLES):
                v = row.get(f"H{i+1}", "")
                cell = ws.cell(row=r, column=4 + i + col_off, value=v if v != "" else None)
                cell.alignment = Alignment(horizontal="center")
                cell.border = bd
            cell_f = ws.cell(row=r, column=4 + HOLES + col_off, value=row.get("Front9", ""))
            cell_f.alignment = Alignment(horizontal="center")
            cell_f.border = bd
            cell_b = ws.cell(row=r, column=5 + HOLES + col_off, value=row.get("Back9", ""))
            cell_b.alignment = Alignment(horizontal="center")
            cell_b.border = bd
            units = payout_units.get(name, 0.0)
            amount = payout_amounts.get(name, 0.0)
            cell_units = ws.cell(row=r, column=6 + HOLES + col_off, value=units if units != 0 else None)
            cell_units.border = bd
            cell_amt = ws.cell(row=r, column=7 + HOLES + col_off, value=round(amount, 2) if amount != 0 else None)
            cell_amt.number_format = '$#,##0.00'
            cell_amt.border = bd

        # highlight birdies/eagles (match GUI colors)
        birdie_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
        eagle_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        for rr in range(start_row + 3, start_row + 3 + len(write_order)):
            for hi in range(HOLES):
                cell = ws.cell(row=rr, column=4 + hi + col_off)
                if cell.value is None:
                    continue
                try:
                    sv = int(cell.value)
                except Exception:
                    continue
                par_for_hole = pars[hi]
                if sv == par_for_hole - 1:
                    cell.fill = birdie_fill
                elif sv <= par_for_hole - 2:
                    cell.fill = eagle_fill
                # ensure hole cells have borders (in case coloring overwrote it visually)
                try:
                    cell.border = bd
                except Exception:
                    pass

        # Summary section (Skins per hole)
        summary_row = start_row + 3 + len(write_order) + 2
        ws.cell(row=summary_row, column=1 + col_off, value="Skins Summary").font = Font(bold=True)
        ws.cell(row=summary_row, column=1 + col_off).border = bd
        summary_row += 1
        ws.cell(row=summary_row, column=1 + col_off, value="Hole").font = Font(bold=True)
        ws.cell(row=summary_row, column=1 + col_off).border = bd
        ws.cell(row=summary_row, column=2 + col_off, value="Result").font = Font(bold=True)
        ws.cell(row=summary_row, column=2 + col_off).border = bd
        ws.cell(row=summary_row, column=3 + col_off, value="Units Awarded").font = Font(bold=True)
        ws.cell(row=summary_row, column=3 + col_off).border = bd
        summary_row += 1
        for hr in hole_results:
            bonus_text = self._format_bonus_summary(hr)
            hole = hr.get("hole", "")
            if hr.get("sole_winner"):
                text = f"{hr['sole_winner']}{bonus_text}"
                units = hr["units_paid"]
            elif hr.get("tied"):
                rreason = hr.get("reason")
                if rreason == REASON_CARRY:
                    text = hr.get("reason_text", f"Tie ({', '.join(hr.get('tied', []))}) - carryover")
                    units = 0
                elif rreason == REASON_SPLIT:
                    text = f"Tie ({', '.join(hr.get('tied', []))}) - split"
                    units = hr.get("units_paid", 0)
                elif rreason == REASON_NO_SCORES:
                    text = hr.get("reason_text", "No scores")
                    units = 0
                else:
                    text = hr.get("reason_text", "No scores")
                    units = 0
                if bonus_text:
                    text = f"{text}{bonus_text}"
            else:
                text = hr.get("reason_text", "No scores")
                if bonus_text:
                    text = f"{text}{bonus_text}"
                units = 0
            c1 = ws.cell(row=summary_row, column=1 + col_off, value=hole)
            c1.border = bd
            c2 = ws.cell(row=summary_row, column=2 + col_off, value=text)
            c2.border = bd
            c3 = ws.cell(row=summary_row, column=3 + col_off, value=units)
            c3.border = bd
            summary_row += 1

        # Payouts table
        summary_row += 1
        ws.cell(row=summary_row, column=1 + col_off, value="Payouts").font = Font(bold=True)
        ws.cell(row=summary_row, column=1 + col_off).border = bd
        summary_row += 1
        ws.cell(row=summary_row, column=1 + col_off, value="Name").font = Font(bold=True)
        ws.cell(row=summary_row, column=1 + col_off).border = bd
        ws.cell(row=summary_row, column=2 + col_off, value="Units").font = Font(bold=True)
        ws.cell(row=summary_row, column=2 + col_off).border = bd
        ws.cell(row=summary_row, column=3 + col_off, value="Amount$").font = Font(bold=True)
        ws.cell(row=summary_row, column=3 + col_off).border = bd
        summary_row += 1
        for name in participants["Name"].tolist():
            units = payout_units.get(name, 0.0)
            amount = payout_amounts.get(name, 0.0)
            c1 = ws.cell(row=summary_row, column=1 + col_off, value=name)
            c1.border = bd
            c2 = ws.cell(row=summary_row, column=2 + col_off, value=round(units, 3) if units != 0 else None)
            c2.border = bd
            c3 = ws.cell(row=summary_row, column=3 + col_off, value=round(amount, 2) if amount != 0 else None)
            c3.number_format = '$#,##0.00'
            c3.border = bd
            summary_row += 1
        if carryover_remaining:
            ws.cell(row=summary_row, column=1 + col_off, value=f"Carryover units remaining after 18: {carryover_remaining}").font = Font(bold=True)
            ws.cell(row=summary_row, column=1 + col_off).border = bd
            summary_row += 1

        # Export Summary sheet (settings + per-player counts)
        try:
            summary = wb.create_sheet(title="Export Summary")
            summary.cell(row=1, column=1 + col_off, value="Setting").font = Font(bold=True)
            summary.cell(row=1, column=2 + col_off, value="Value").font = Font(bold=True)
            r = 2

            def srow(k, v):
                nonlocal r
                summary.cell(row=r, column=1 + col_off, value=k)
                summary.cell(row=r, column=2 + col_off, value=v)
                # add border to summary entries
                try:
                    summary.cell(row=r, column=1 + col_off).border = bd
                    summary.cell(row=r, column=2 + col_off).border = bd
                except Exception:
                    pass
                r += 1

            srow("Course", self.course_var.get().strip())
            try:
                srow("Date", self.date_entry.get_date().strftime("%Y-%m-%d"))
            except Exception:
                srow("Date", self.date_var.get().strip())
            srow("Per-skin $", float(per_skin))
            if tp != "":
                try:
                    srow("Total Purse $", float(tp))
                except Exception:
                    srow("Total Purse $", tp)
            srow("Carryover Enabled", str(self.carryover_var.get()))
            srow("Use Net Scores", str(self.use_net_scores.get()))
            srow("Bonuses Enabled", str(self.bonus_enabled_var.get()))
            srow("Split Ties", str(self.split_ties.get()))

            # simple per-player summary
            participants_df = df[df.get("Included") == True]
            player_names = participants_df.get("Name").tolist()
            units_map = results.get("payout_map_units", {})
            birdie_map = {n: 0 for n in player_names}
            eagle_map = {n: 0 for n in player_names}
            for _, prow in participants_df.iterrows():
                name = prow.get("Name")
                for i in range(HOLES):
                    try:
                        val = prow.get(f"H{i+1}")
                        if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
                            continue
                        score = int(val)
                    except Exception:
                        continue
                    parv = int(pars[i]) if i < len(pars) else 4
                    if score == parv - 1:
                        birdie_map[name] += 1
                    elif score <= parv - 2:
                        eagle_map[name] += 1

            r += 1
            summary.cell(row=r, column=1 + col_off, value="Player").font = Font(bold=True)
            summary.cell(row=r, column=2 + col_off, value="Total Units").font = Font(bold=True)
            summary.cell(row=r, column=3 + col_off, value="Birdies").font = Font(bold=True)
            summary.cell(row=r, column=4 + col_off, value="Eagles").font = Font(bold=True)
            summary.cell(row=r, column=5 + col_off, value="Amount$").font = Font(bold=True)
            r += 1

            total_paid = 0.0
            for name in player_names:
                u = units_map.get(name, 0.0)
                amt = payout_amounts.get(name, 0.0)
                summary.cell(row=r, column=1 + col_off, value=name)
                summary.cell(row=r, column=2 + col_off, value=float(u) if u != 0 else None)
                summary.cell(row=r, column=3 + col_off, value=birdie_map.get(name, 0) if birdie_map.get(name, 0) != 0 else None)
                summary.cell(row=r, column=4 + col_off, value=eagle_map.get(name, 0) if eagle_map.get(name, 0) != 0 else None)
                cell_amt = summary.cell(row=r, column=5 + col_off, value=round(amt or 0.0, 2))
                cell_amt.number_format = '$#,##0.00'
                total_paid += float(amt or 0.0)
                r += 1

            if total_paid:
                summary.cell(row=r, column=1 + col_off, value="Total Paid").font = Font(bold=True)
                cell_total = summary.cell(row=r, column=5 + col_off, value=round(total_paid, 2))
                cell_total.number_format = '$#,##0.00'
                cell_total.font = Font(bold=True)

        except Exception:
            # safe to continue even if summary sheet creation failed
            pass

        for c in range(1, 8 + HOLES):
            ws.column_dimensions[get_column_letter(c + col_off)].width = 14

        safe_course = ''.join(c for c in (self.course_var.get().strip() or 'course') if c.isalnum() or c in (' ', '_', '-')).replace(' ', '_')
        default_name = f"BigBoySkins_{safe_course}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name,
                                            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        wb.save(path)
        messagebox.showinfo("Exported", f"Report exported to {path}")
        

    def import_from_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active

            # read optional Export Summary
            try:
                summary_sheet = None
                for name in wb.sheetnames:
                    if name.strip().lower() in ("export summary", "export_summary"):
                        summary_sheet = wb[name]
                        break
                if summary_sheet is not None:
                    def _to_bool(v):
                        if isinstance(v, bool):
                            return v
                        if v is None:
                            return False
                        s = str(v).strip().lower()
                        return s in ("true", "1", "yes")
                    for row in summary_sheet.iter_rows(values_only=True):
                        if not row:
                            continue
                        # find first non-empty cell in the row (Export Summary may be shifted)
                        first_cell = None
                        first_idx = None
                        for j, cellval in enumerate(row):
                            try:
                                if cellval is not None and str(cellval).strip() != "":
                                    first_cell = cellval
                                    first_idx = j
                                    break
                            except Exception:
                                continue
                        if first_cell is None:
                            continue
                        key = str(first_cell).strip().lower()
                        val = row[first_idx + 1] if (first_idx is not None and len(row) > first_idx + 1) else None
                        try:
                            if "per-skin" in key:
                                self.per_skin_var.set(str(val))
                            elif "total purse" in key:
                                self.total_purse_var.set(str(val))
                            elif "carryover" in key:
                                self.carryover_var.set(_to_bool(val))
                            elif "use net" in key:
                                self.use_net_scores.set(_to_bool(val))
                            elif "split ties" in key:
                                self.split_ties.set(_to_bool(val))
                            elif key == "course":
                                self.course_var.set(str(val))
                            elif key == "date" and val is not None:
                                try:
                                    if hasattr(val, 'strftime'):
                                        self.date_entry.set_date(val)
                                    else:
                                        self.date_entry.set_date(datetime.strptime(str(val), "%Y-%m-%d").date())
                                except Exception:
                                    self.date_var.set(str(val))
                        except Exception:
                            pass
            except Exception:
                pass

            # find header row with "Name" (search first 50 rows × first 10 columns)
            def _norm(v):
                if v is None:
                    return ""
                s = str(v)
                # remove common invisible/formatting chars
                for ch in ('\ufeff', '\u200b', '\u2060', '\u00a0'):
                    s = s.replace(ch, '')
                return s.strip().lower()

            hr = None
            header_col = 1
            for r in range(1, 51):
                for c in range(1, 11):
                    try:
                        val = ws.cell(row=r, column=c).value
                    except Exception:
                        val = None
                    if _norm(val).find('name') != -1 or _norm(val).find('player') != -1:
                        hr = r
                        header_col = c
                        break
                if hr is not None:
                    break

            if hr is None:
                # write a small diagnostic snapshot to help debugging
                try:
                    log_path = os.path.join(os.getcwd(), 'import_error.log')
                    with open(log_path, 'w', encoding='utf-8') as lf:
                        lf.write('Could not find header row. Sheet snapshot (first 20 rows x 10 cols):\n')
                        for rr in range(1, 21):
                            row_vals = []
                            for cc in range(1, 11):
                                try:
                                    row_vals.append(str(ws.cell(row=rr, column=cc).value))
                                except Exception:
                                    row_vals.append('')
                            lf.write(f'Row {rr}: ' + '\t'.join(row_vals) + '\n')
                    messagebox.showerror('Import failed', f'Could not find header row in spreadsheet. See {log_path}')
                except Exception:
                    messagebox.showerror('Import failed', 'Could not find header row in spreadsheet')
                return

            rows = []
            r = hr + 1

            # Detect optional Par row
            first_label = ws.cell(row=r, column=header_col).value
            if isinstance(first_label, str) and first_label.strip().lower() == "par":
                for i in range(HOLES):
                    try:
                        val = ws.cell(row=r, column=header_col + 3 + i).value
                        if val is None:
                            self.par_vars[i].set("4")
                        else:
                            self.par_vars[i].set(str(int(val)) if isinstance(val, (int, float)) else str(val))
                    except Exception:
                        self.par_vars[i].set("4")
                r += 1

            # Detect optional Stroke Index row
            second_label = ws.cell(row=r, column=header_col).value
            if isinstance(second_label, str) and second_label.strip().lower() in ("stroke index", "stroke_index", "si"):
                for i in range(HOLES):
                    try:
                        val = ws.cell(row=r, column=header_col + 3 + i).value
                        if val is None:
                            self.stroke_index_vars[i].set(str(i+1))
                        else:
                            self.stroke_index_vars[i].set(str(val))
                    except Exception:
                        self.stroke_index_vars[i].set(str(i+1))
                r += 1

            def _is_bool_like(val):
                if isinstance(val, bool):
                    return True
                if isinstance(val, (int, float)) and val in (0, 1):
                    return True
                try:
                    s = str(val).strip().lower()
                    if s in ("true", "false", "1", "0", "yes", "no"):
                        return True
                except Exception:
                    pass
                return False

            hole_col_start = header_col + 3
            def _row_has_scores(ws_obj, row_idx):
                for i in range(HOLES):
                    v = ws_obj.cell(row=row_idx, column=hole_col_start + i).value
                    if v is None:
                        continue
                    if isinstance(v, (int, float)):
                        return True
                    if str(v).strip() != "":
                        return True
                return False

            while True:
                name = ws.cell(row=r, column=header_col).value
                if name is None:
                    break
                if isinstance(name, str) and name.strip().lower() in ("skins summary", "payouts"):
                    break
                included_cell = ws.cell(row=r, column=header_col + 2).value
                hcp_cell = ws.cell(row=r, column=header_col + 1).value
                if not (_is_bool_like(included_cell) or _row_has_scores(ws, r) or (hcp_cell is not None and str(hcp_cell).strip() != "")):
                    break
                included = included_cell
                entry = {"Name": str(name), "Included": bool(included)}
                hcp = ws.cell(row=r, column=header_col + 1).value
                entry["Handicap"] = hcp if hcp is not None else "0"
                for i in range(HOLES):
                    entry[f"H{i+1}"] = ws.cell(row=r, column=hole_col_start + i).value
                entry["Front9"] = ws.cell(row=r, column=hole_col_start + HOLES).value
                entry["Back9"] = ws.cell(row=r, column=hole_col_start + HOLES + 1).value
                rows.append(entry)
                r += 1

            for p in self.players:
                p.destroy()
            self.players = []
            for i, row in enumerate(rows):
                pr = PlayerRow(self.player_inner, i, self)
                pr.load_from_dict(row)
                self.players.append(pr)
            while len(self.players) < 2:
                self.add_player()
            # adjust height after import so all imported rows are visible
            try:
                self._adjust_height()
            except Exception:
                pass

            # import metadata (respect detected header column)
            try:
                course = ws.cell(row=2, column=header_col + 1).value
                date_val = ws.cell(row=3, column=header_col + 1).value
                if course:
                    self.course_var.set(str(course))
                if isinstance(date_val, datetime):
                    try:
                        self.date_entry.set_date(date_val.date())
                    except:
                        pass
                else:
                    try:
                        self.date_entry.set_date(datetime.strptime(str(date_val), "%Y-%m-%d").date())
                    except:
                        pass
            except:
                pass

            messagebox.showinfo("Imported", "Contest imported successfully")
        except Exception as e:
            messagebox.showerror("Import error", f"Failed to import: {e}")


def main():
    root = tk.Tk()
    try:
        icon_path = get_app_icon_path()
        if os.path.exists(icon_path):
            icon_img = tk.PhotoImage(file=icon_path)
            root.iconphoto(True, icon_img)
            # Keep a reference to avoid image garbage-collection in Tk.
            root._icon_img = icon_img
    except Exception:
        pass
    root.geometry("2040x700")
    app = BigBoySkinsApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()